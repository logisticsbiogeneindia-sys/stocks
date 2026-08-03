
import io, zipfile
from copy import copy
import streamlit as st
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title="Excel Splitter",layout="wide")
st.title("📄 Excel Workbook Splitter")

up=st.file_uploader("Upload workbook",type=["xlsx"])
if up:
    wb=load_workbook(up)
    sheet=st.selectbox("Worksheet",wb.sheetnames)
    ws=wb[sheet]
    headers=[str(c.value).strip() if c.value else "" for c in ws[1]]
    col=st.selectbox("Split Column",headers)
    mode=st.radio("Mode",["Worksheets in one workbook","Separate Excel files"])
    ignore=st.checkbox("Ignore blank values",True)
    if st.button("Split"):
        idx=headers.index(col)+1
        prog=st.progress(0)
        rows=list(ws.iter_rows(min_row=2))
        total=max(1,len(rows))
        def clean(v):
            if v is None or str(v).strip()=="":
                return None
            n=str(v).strip()
            for ch in '\\/*[]:?':
                n=n.replace(ch,'_')
            return n[:31]
        if mode.startswith("Worksheets"):
            out=Workbook()
            out.remove(out.active)
            sheets={}
            for i,row in enumerate(rows,1):
                name=clean(row[idx-1].value)
                if name is None:
                    if ignore:
                        prog.progress(i/total); continue
                    name="Blank"
                if name not in sheets:
                    sh=out.create_sheet(name)
                    sheets[name]=sh
                    for c,h in enumerate(ws[1],1):
                        sh.cell(1,c).value=h.value
                sh=sheets[name]
                nr=sh.max_row+1
                for c,cell in enumerate(row,1):
                    sh.cell(nr,c).value=cell.value
                prog.progress(i/total)
            bio=io.BytesIO(); out.save(bio); bio.seek(0)
            st.success(f"Created {len(sheets)} worksheets")
            st.download_button("Download Workbook",bio,"Split_Workbook.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            groups={}
            for i,row in enumerate(rows,1):
                name=clean(row[idx-1].value)
                if name is None:
                    if ignore:
                        prog.progress(i/total); continue
                    name="Blank"
                groups.setdefault(name,[]).append(row)
                prog.progress(i/total)
            zbio=io.BytesIO()
            with zipfile.ZipFile(zbio,"w",zipfile.ZIP_DEFLATED) as z:
                for name,grows in groups.items():
                    ow=Workbook(); os=ow.active; os.title=name
                    for c,h in enumerate(ws[1],1):
                        os.cell(1,c).value=h.value
                    for r,row in enumerate(grows,2):
                        for c,cell in enumerate(row,1):
                            os.cell(r,c).value=cell.value
                    x=io.BytesIO(); ow.save(x)
                    z.writestr(f"{name}.xlsx",x.getvalue())
            zbio.seek(0)
            st.success(f"Created {len(groups)} files")
            st.download_button("Download ZIP",zbio,"Split_Files.zip","application/zip")
